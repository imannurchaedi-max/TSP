"""
Read column headings and sample data from the 'BARCODE OUTBOUND WRM' tab.
"""
import openpyxl
import sys
import os

XLSX_PATH = os.path.join(os.path.dirname(__file__), "TSP MODUL.xlsx")
TARGET_SHEET = "BARCODE OUTBOUND WRM"

def main():
    print(f"Opening: {XLSX_PATH}")
    print(f"File size: {os.path.getsize(XLSX_PATH) / 1024 / 1024:.1f} MB\n")

    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)

    # Find the target sheet (handle trailing spaces)
    ws = None
    actual_name = None
    for name in wb.sheetnames:
        if name.strip() == TARGET_SHEET:
            ws = wb[name]
            actual_name = name
            break

    if ws is None:
        print(f"ERROR: Sheet '{TARGET_SHEET}' not found!")
        print(f"Available sheets: {wb.sheetnames}")
        wb.close()
        sys.exit(1)

    print(f"=== SHEET: \"{actual_name}\" ===\n")

    # Read all rows (header + data) up to max 10 rows
    all_rows = []
    for i, row in enumerate(ws.iter_rows(values_only=False)):
        if i >= 12:  # header + 10 data rows + 1 buffer
            break
        all_rows.append(row)

    if not all_rows:
        print("Sheet is empty!")
        wb.close()
        return

    # Header row
    header_row = all_rows[0]
    print("=== COLUMN HEADINGS (Row 1) ===")
    headers = []
    for cell in header_row:
        if cell.value is not None:
            val_str = str(cell.value).strip()
            headers.append((cell.column, cell.column_letter, val_str))
            print(f"  Col {cell.column:3d} ({cell.column_letter:>3s}): \"{val_str}\"")

    print(f"\nTotal columns with data: {len(headers)}")
    print()

    # Sample data rows
    print("=== SAMPLE DATA (First 5 data rows) ===")
    data_rows = all_rows[1:6]
    for row_offset, row in enumerate(data_rows, 2):
        print(f"\n--- Row {row_offset} ---")
        for col_num, col_letter, col_name in headers:
            for cell in row:
                if cell.column == col_num:
                    if cell.value is not None:
                        print(f"  {col_name}: {repr(cell.value)}")
                    break

    # Count approximate total rows (read a bit more)
    row_count = len(all_rows) - 1  # minus header
    print(f"\n(Loaded {row_count} data rows in preview)")

    wb.close()
    print("\nDone.")

if __name__ == "__main__":
    main()
